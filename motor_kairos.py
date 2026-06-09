import sys
import io
import os
import json
import base64
import asyncio
import threading
from urllib.parse import urlparse, parse_qs
from datetime import datetime, timedelta
from http.server import BaseHTTPRequestHandler, HTTPServer

# MÁGICA DA NUVEM: Ensina o Python a achar o navegador baixado pelo build.sh na Render
if os.environ.get("RENDER"):
    os.environ["PLAYWRIGHT_BROWSERS_PATH"] = "0"

import pandas as pd
from playwright.async_api import async_playwright
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import ColorScaleRule

if sys.stdout is not None:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace', line_buffering=True)
else:
    sys.stdout = open(os.devnull, 'w')

if sys.stderr is not None:
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace', line_buffering=True)
else:
    sys.stderr = open(os.devnull, 'w')

# =============================================
# CONFIGURAÇÕES DA FECHADURA ANTI-PIRATARIA
# =============================================
import urllib.request
import urllib.error

def verificar_token_seguranca(token):
    """
    Trava Anti-Pirataria: Validação Remota de Licença via API Central.
    """
    if not token:
        return False
        
    # Backdoor seguro para que você possa continuar testando agora mesmo
    if token == "KAIROS2026":
        return True
        
    URL_API_KAIROS = f"https://sua-api-kairos.com/validar_token?token={token}"
    
    try:
        # Timeout curto de 5 segundos exigido pela arquitetura de segurança
        req = urllib.request.Request(URL_API_KAIROS)
        with urllib.request.urlopen(req, timeout=5.0) as response:
            if response.getcode() == 200:
                body = response.read().decode('utf-8')
                dados = json.loads(body)
                # Se o status vier ativo da API, libera o acesso
                if dados.get("status") == "ativo":
                    return True
    except Exception as e:
        print(f"⚠️ Proteção Ativada: Falha de rede ou licença recusada ao contatar API: {e}")
        # Bloqueio estrito: Se der timeout, erro de rede, ou a API cair, RETORNA FALSO!
        pass
        
    return False

# =============================================
# CONFIGURAÇÕES DO SCRAPER
# =============================================
DURACAO = {
    "Futebol": 105, "Basquete": 150, "Tenis": 120, "Hoquei": 150,
    "Futebol Americano": 190, "Beisebol": 180
}

LIGAS_PRIORITARIAS = [
    "premier league", "serie a", "brasileirao", "brasileirão",
    "copa do brasil", "libertadores", "sul-americana",
    "nba", "euroleague", "euroliga", "nbb",
    "atp", "wta", "itf", "challenger",
    "nhl", "shl", "del", "segunda divisao", "serie b",
    "nfl", "mlb"
]

URLS_ESPORTES = {
    "Futebol":  "https://www.flashscore.com.br/futebol/",
    "Basquete": "https://www.flashscore.com.br/basquete/",
    "Tenis":    "https://www.flashscore.com.br/tenis/",
    "Hoquei":   "https://www.flashscore.com.br/hoquei/", 
    "Futebol Americano": "https://www.flashscore.com.br/futebol-americano/",
    "Beisebol": "https://www.flashscore.com.br/beisebol/"
}

o_loop = None # Guardará o loop do asyncio

# =============================================
# SCRAPING ROBUSTO E OTIMIZADO PARA NUVEM
# =============================================
async def rolar_pagina_completa(page):
    print("   📜 Rolando página completa de forma acelerada...")
    altura_anterior = 0
    tentativas_sem_mudanca = 0
    while tentativas_sem_mudanca < 3:
        altura_atual = await page.evaluate("document.body.scrollHeight")
        if altura_atual == altura_anterior: tentativas_sem_mudanca += 1
        else: tentativas_sem_mudanca = 0
        altura_anterior = altura_atual
        await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        await page.wait_for_timeout(300) # Reduzido de 1500 para 300ms
    await page.evaluate("window.scrollTo(0, 0)")
    await page.wait_for_timeout(200)

async def navegar_para_data(page, data_alvo: datetime):
    hoje = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    alvo = data_alvo.replace(hour=0, minute=0, second=0, microsecond=0)
    diff_dias = (alvo - hoje).days
    if diff_dias == 0:
        return
    seta = "button[data-day-picker-arrow='next']" if diff_dias > 0 else "button[data-day-picker-arrow='prev']"
    for i in range(abs(diff_dias)):
        try:
            await page.click(seta, timeout=3000)
            await page.wait_for_timeout(500)
        except Exception as e:
            break

async def extrair_texto_esporte(page, esporte, url_base, data: datetime):
    print(f"   🌐 Acessando {esporte} → {url_base}")
    try:
        await page.goto(url_base, wait_until="domcontentloaded", timeout=60000)
        await page.wait_for_timeout(1000) # Reduzido de 5000 para 1000ms
        try:
            await page.click("button#onetrust-accept-btn-handler", timeout=2000)
            await page.wait_for_timeout(500)
        except: pass
        
        await navegar_para_data(page, data)
        await page.wait_for_timeout(1000)
        
        hoje = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        alvo = data.replace(hour=0, minute=0, second=0, microsecond=0)
        if alvo > hoje:
            for seletor_aba in ["button:has-text('PRÓXIMOS')", "text=PRÓXIMOS", "button:has-text('TODOS')", "text=TODOS"]:
                try:
                    await page.click(seletor_aba, timeout=2000)
                    await page.wait_for_timeout(1000)
                    break
                except: continue
        elif alvo < hoje:
            for seletor_aba in ["button:has-text('ENCERRADOS')", "text=ENCERRADOS"]:
                try:
                    await page.click(seletor_aba, timeout=2000)
                    await page.wait_for_timeout(1000)
                    break
                except: continue
        else:
            for seletor_aba in ["button:has-text('TODOS')", "text=TODOS"]:
                try:
                    await page.click(seletor_aba, timeout=2000)
                    await page.wait_for_timeout(1000)
                    break
                except: continue
                
        await rolar_pagina_completa(page)
        texto = await page.evaluate("document.body.innerText")
        print(f"   ✅ {esporte} coletado!")
        return texto
    except Exception as e:
        print(f"   ⚠️ Erro em {esporte}: {e}")
        return ""

# Função interceptadora para bloquear peso inútil
async def interceptar_rota(route):
    if route.request.resource_type in ["image", "stylesheet", "media", "font", "other"]:
        await route.abort()
    else:
        await route.continue_()

async def extrair_jogos_flashscore(data: datetime = None):
    if data is None: data = datetime.now()
    dados_por_esporte = {}
    
    for esporte, url in URLS_ESPORTES.items():
        async with async_playwright() as p:
            # --disable-dev-shm-usage é OBRIGATÓRIO no Docker/Render para não estourar a memória compartilhada
            browser = await p.chromium.launch(headless=True, args=["--start-maximized", "--disable-dev-shm-usage", "--no-sandbox"])
            context = await browser.new_context(viewport={"width": 1366, "height": 768}, user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
            
            # OTIMIZAÇÃO EXTREMA: Bloqueia carregamento de imagens, vídeos, fontes e estilos CSS
            await context.route("**/*", interceptar_rota)
            
            page = await context.new_page()
            texto = await extrair_texto_esporte(page, esporte, url, data)
            if texto: dados_por_esporte[esporte] = texto
            
            await browser.close() # Mata o navegador inteiramente após cada esporte, esvaziando a RAM a zero
            
    return dados_por_esporte

# =============================================
# EXTRAÇÃO E CÁLCULOS
# =============================================
def extrair_jogos_esporte_direto(esporte, texto):
    import re
    SKIP = {
        'PREVIEW', 'SRF', 'Classificação', 'Tabela', 'Classificação ao vivo',
        'TODOS', 'AO VIVO', 'ODDS', 'ENCERRADOS', 'PRÓXIMOS', 'Publicidade',
        'LIGAS FIXADAS', 'MINHAS EQUIPES', 'ADICIONAR EQUIPE', 'RANKINGS',
        'FAVORITOS', 'FUTEBOL', 'BASQUETE', 'TÊNIS', 'VÔLEI', 'FUTEBOL AM.',
        'BEISEBOL', 'HANDEBOL', 'ACESSAR', 'RESULTADOS', 'NOTÍCIAS', 'APOSTAS',
        'FIFA', 'PAÍSES', 'Mostrar mais', '-', 'CALENDÁRIO', 'CATEGORIAS', 'TORNEIOS ATUAIS',
    }
    time_re  = re.compile(r'^\d{2}:\d{2}$')
    pais_re  = re.compile(r'^[A-ZÁÉÍÓÚÀÃÕÂÊÎÔÛÇ][A-ZÁÉÍÓÚÀÃÕÂÊÎÔÛÇa-záéíóúàãõâêîôû\s\.\-]+:\s*$')
    hidden_re = re.compile(r'^exibir jogos \(\d+\)$')
    score_re  = re.compile(r'^\d+$')
    date_re   = re.compile(r'^\d{2}/\d{2}')

    linhas = [l.strip() for l in texto.split('\n')]
    jogos = []
    liga_atual = ''
    i = 0
    while i < len(linhas):
        linha = linhas[i]
        if (not linha or linha in SKIP or hidden_re.match(linha) or score_re.match(linha) or date_re.match(linha)):
            i += 1
            continue
        if time_re.match(linha):
            times = []
            j = i + 1
            while j < len(linhas) and len(times) < 2:
                cand = linhas[j].strip()
                if (not cand or cand in SKIP or cand == 'SRF' or score_re.match(cand) or hidden_re.match(cand)):
                    j += 1
                    continue
                if time_re.match(cand) or pais_re.match(cand): break
                times.append(cand)
                j += 1
            if len(times) == 2 and liga_atual:
                jogos.append({'h': linha, 'e': esporte, 'l': liga_atual, 'm': times[0], 'v': times[1]})
            i += 1
            continue
        if pais_re.match(linha):
            i += 1
            continue
        for j in range(i + 1, min(i + 4, len(linhas))):
            prox = linhas[j].strip()
            if not prox: continue
            if pais_re.match(prox) or prox in {'Classificação', 'Tabela', 'Classificação ao vivo'}:
                liga_atual = linha
            break
        i += 1
    return jogos

def extrair_todos_jogos_com_ia(dados_por_esporte):
    todos_jogos = []
    for esporte, texto in dados_por_esporte.items():
        todos_jogos.extend(extrair_jogos_esporte_direto(esporte, texto))
    return todos_jogos

def calcular_heatmap(jogos):
    janelas = [(datetime.strptime("00:00", "%H:%M") + timedelta(minutes=30 * i)).strftime("%H:%M") for i in range(48)]
    heatmap = []
    for janela in janelas:
        h_janela = datetime.strptime(janela, "%H:%M")
        c = {"janela": janela, "Futebol": 0, "Basquete": 0, "Tenis": 0, "Hoquei": 0, "Futebol Americano": 0, "Beisebol": 0, "total": 0}
        for jogo in jogos:
            try:
                h_jogo = datetime.strptime(jogo['h'], "%H:%M")
                e = jogo['e']
                h_fim = h_jogo + timedelta(minutes=DURACAO.get(e, 105))
                if h_jogo <= h_janela < h_fim:
                    if e in c: c[e] += 1
                    c["total"] += 1
            except: continue
        heatmap.append(c)
    return heatmap

def calcular_power_hours(heatmap): 
    return sorted(heatmap, key=lambda x: x['total'], reverse=True)[:3]

def calcular_alertas(jogos, power_hours):
    alertas = []
    janelas_power = [p['janela'] for p in power_hours]
    for jogo in jogos:
        liga = jogo.get('l', '').lower()
        if not any(lp in liga for lp in LIGAS_PRIORITARIAS): continue
        try:
            h_jogo = datetime.strptime(jogo['h'], "%H:%M")
            h_fim = h_jogo + timedelta(minutes=DURACAO.get(jogo['e'], 105))
            for janela in janelas_power:
                h_j = datetime.strptime(janela, "%H:%M")
                if h_jogo <= h_j < h_fim:
                    alertas.append({"janela": janela, "jogo": f"{jogo['m']} x {jogo['v']}", "liga": jogo['l'], "esporte": jogo['e']})
        except: continue
    return alertas

# =============================================
# EXCEL MELHORADO
# =============================================
def gerar_excel(jogos, heatmap, power_hours, alertas, data: datetime = None):
    if data is None: data = datetime.now()
    filename = f"Analise_KAIROS_{data.strftime('%d_%m_%Y')}.xlsx"
    
    resumo = {
        "Info": [
            "📅 Data Analisada", "🎮 Total de Jogos", 
            "⚽ Futebol", "🏀 Basquete", "🎾 Tênis", "🏒 Hóquei", "🏈 Fut. Americano", "⚾ Beisebol", 
            "🔥 Pico 1", "🔥 Pico 2", "🔥 Pico 3"
        ],
        "Valor": [
            data.strftime('%d/%m/%Y'), len(jogos), 
            sum(1 for j in jogos if j['e'] == 'Futebol'), 
            sum(1 for j in jogos if j['e'] == 'Basquete'), 
            sum(1 for j in jogos if j['e'] == 'Tenis'), 
            sum(1 for j in jogos if j['e'] == 'Hoquei'), 
            sum(1 for j in jogos if j['e'] == 'Futebol Americano'), 
            sum(1 for j in jogos if j['e'] == 'Beisebol'), 
            f"{power_hours[0]['janela']} ({power_hours[0]['total']} jogos)" if len(power_hours) > 0 else "-", 
            f"{power_hours[1]['janela']} ({power_hours[1]['total']} jogos)" if len(power_hours) > 1 else "-", 
            f"{power_hours[2]['janela']} ({power_hours[2]['total']} jogos)" if len(power_hours) > 2 else "-"
        ]
    }
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        pd.DataFrame(resumo).to_excel(writer, sheet_name='Resumo', index=False)
        
        df_heat = pd.DataFrame(heatmap)
        df_heat.columns = ["Janela", "Futebol", "Basquete", "Tênis", "Hóquei", "Fut. Americano", "Beisebol", "Total"]
        df_heat.to_excel(writer, sheet_name='Heatmap', index=False)
        
        if alertas:
            df_al = pd.DataFrame(alertas)
            df_al.columns = ["Janela", "Jogo", "Liga", "Esporte"]
            df_al.to_excel(writer, sheet_name='Jogos Power Hour', index=False)
        else:
            pd.DataFrame({"Aviso": ["Nenhum jogo prioritário"]}).to_excel(writer, sheet_name='Jogos Power Hour', index=False)
            
        # --- APLICAR ESTILOS NO EXCEL (HEATMAP) ---
        workbook = writer.book
        worksheet = writer.sheets['Heatmap']
        
        # Estilo do Cabeçalho
        header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid") # Azul escuro
        header_font = Font(color="FFFFFF", bold=True) # Branco
        center_align = Alignment(horizontal="center", vertical="center")
        
        for col_num in range(1, 9):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            # Ajustar largura da coluna
            col_letter = worksheet.cell(row=1, column=col_num).column_letter
            worksheet.column_dimensions[col_letter].width = 15
            
        # Aplicar formatação e alinhamento para os dados
        for row in range(2, len(df_heat) + 2):
            for col in range(1, 9):
                worksheet.cell(row=row, column=col).alignment = center_align

        # Regra de Cores Clássica do Excel com Piso Mínimo (ColorScaleRule)
        colunas_esportes = {
            'B': 'Futebol', 'C': 'Basquete', 'D': 'Tênis', 
            'E': 'Hóquei', 'F': 'Fut. Americano', 'G': 'Beisebol', 'H': 'Total'
        }
        
        for col_letter, col_name in colunas_esportes.items():
            max_val = df_heat[col_name].max()
            
            piso = 30 if col_name == 'Total' else 15
            pico_escala = max(max_val, piso)
            
            regra = ColorScaleRule(
                start_type='num', start_value=0, start_color='00B050',
                mid_type='num', mid_value=pico_escala / 2.0, mid_color='FFEB84',
                end_type='num', end_value=pico_escala, end_color='FF0000'
            )
            worksheet.conditional_formatting.add(f'{col_letter}2:{col_letter}{len(df_heat) + 1}', regra)
            
    return filename

# =============================================
# SERVIDOR CLOUD API (FASTAPI)
# =============================================
from fastapi import FastAPI, HTTPException, Query, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
import uvicorn
import uuid

app = FastAPI(title="Motor Kairós SaaS API")

# Dicionário em memória para gerenciar o Polling (Tickets)
tarefas_varredura = {}

# Libera o bloqueio de CORS para o site da Vercel
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], # Permitir de qualquer origem (Vercel ou testes locais)
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/")
def health_check():
    return {"status": "online", "mensagem": "Motor Kairós operando na Nuvem ☁️"}

async def processar_varredura_background(task_id: str, dt: datetime, data_cmd: str):
    try:
        tarefas_varredura[task_id] = {"status": "processando", "mensagem": "Raspando os esportes..."}
        dados_brutos = await extrair_jogos_flashscore(dt)
        if not dados_brutos:
            tarefas_varredura[task_id] = {"status": "erro", "mensagem": "Nenhum dado encontrado no Flashscore."}
            return

        jogos = extrair_todos_jogos_com_ia(dados_brutos)
        if not jogos:
            tarefas_varredura[task_id] = {"status": "erro", "mensagem": "Nenhum jogo estruturado extraído."}
            return

        heatmap = calcular_heatmap(jogos)
        power_hours = calcular_power_hours(heatmap)
        alertas = calcular_alertas(jogos, power_hours)
        
        arquivo_excel = gerar_excel(jogos, heatmap, power_hours, alertas, dt)
        
        with open(arquivo_excel, "rb") as f:
            bytes_planilha = f.read()
        planilha_base64 = base64.b64encode(bytes_planilha).decode('utf-8')
        os.remove(arquivo_excel)
        
        tarefas_varredura[task_id] = {
            "status": "concluido",
            "dados_painel": {"total_jogos": len(jogos)},
            "planilha_nome": arquivo_excel,
            "planilha_base64": planilha_base64
        }
        print(f"✅ Tarefa {task_id} concluída com sucesso!")
    except Exception as e:
        print(f"❌ Erro na tarefa {task_id}: {e}")
        tarefas_varredura[task_id] = {"status": "erro", "mensagem": str(e)}

@app.get("/analisar")
async def endpoint_analisar(background_tasks: BackgroundTasks, data: str = Query("hoje"), token: str = Query("")):
    print(f"📡 Requisição recebida. Data: {data}")
    
    if not verificar_token_seguranca(token):
        print("❌ Bloqueio: Licença inválida.")
        raise HTTPException(status_code=401, detail="Assinatura Kairós inativa ou inválida. Renove seu acesso.")
    
    data_cmd = data.lower().strip()
    if data_cmd == "hoje": dt = datetime.now()
    elif data_cmd in ["amanha", "amanhã"]: dt = datetime.now() + timedelta(days=1)
    else:
        try: dt = datetime.strptime(data_cmd, "%d/%m/%Y")
        except: dt = datetime.now()

    task_id = str(uuid.uuid4())
    background_tasks.add_task(processar_varredura_background, task_id, dt, data_cmd)
    
    return {
        "status": "iniciado",
        "task_id": task_id,
        "mensagem": "Varredura iniciada em nuvem. Acompanhe o status."
    }

@app.get("/status_varredura")
async def endpoint_status(task_id: str = Query(...)):
    tarefa = tarefas_varredura.get(task_id)
    if not tarefa:
        raise HTTPException(status_code=404, detail="Tarefa não encontrada ou expirada.")
    return tarefa

if __name__ == "__main__":
    # Inicia o servidor Web Uvicorn (usado pela Render)
    port = int(os.environ.get("PORT", 8000))
    print(f"🚀 Iniciando servidor FastAPI Kairós na porta {port}...")
    uvicorn.run("motor_kairos:app", host="0.0.0.0", port=port)
